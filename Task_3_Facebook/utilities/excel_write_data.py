from openpyxl import Workbook
from openpyxl import load_workbook
import os
import glob

def createFile(fileDirectory, fileName, sheetName):
    file = Workbook()
    file_sheet = file.active
    file_sheet.title = sheetName
    file_sheet.cell(row=1, column=1).value = 'Demand'
    file_sheet.cell(row=1, column=2).value = 'Status'
    file_sheet.cell(row=1, column=3).value = 'Ticket_ID'
    file_sheet.cell(row=1, column=4).value = 'Questions'
    file_sheet.cell(row=1, column=5).value = 'Notes'
    file.save(fileDirectory + fileName)
    file.close()

def writeFileRows(data1, data2, data3, data4, data5):
    path = os.getcwd()
    ticket_path = path + r'\tickets\*'
    list_of_files = glob.glob(ticket_path)
    latest_file = max(list_of_files, key=os.path.getctime)
    file = load_workbook(latest_file)
    file_sheet = file.active
    maximum = file_sheet.max_row
    file_sheet.cell(row=maximum + 1, column=1).value = data1
    file_sheet.cell(row=maximum + 1, column=2).value = data2
    file_sheet.cell(row=maximum + 1, column=3).value = data3
    file_sheet.cell(row=maximum + 1, column=4).value = data4
    file_sheet.cell(row=maximum + 1, column=5).value = data5
    file.save(latest_file)
    file.close()
