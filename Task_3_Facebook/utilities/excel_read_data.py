from openpyxl import load_workbook


def searchDataRow(searchItem, fileName, sheetName="Sheet1"):

    excelFile =load_workbook(fileName)
    excelFile_sheet = excelFile[sheetName]

    no_rows = excelFile_sheet.max_row

    for row in range(1, no_rows+1):
        if excelFile_sheet.cell(row=row, column=1).value == searchItem:
            correspondingValue = excelFile_sheet.cell(row=row, column=2).value
            return correspondingValue

def searchAppFile(searchItem, fileName, rowCount):

    excelFile = load_workbook(fileName)
    excelFile_sheet = excelFile.active

    no_cols = excelFile_sheet.max_column

    for col in range(1, no_cols+1):
        if excelFile_sheet.cell(row=1, column=col).value == searchItem:
            correspondingValue = excelFile_sheet.cell(row=rowCount, column=col).value
            return correspondingValue

def readFile(chatFile, chatFile_row, chatFile_col):

    chatFile = load_workbook(chatFile)
    chatFile_sheet = chatFile.active

    if chatFile_sheet.cell(row=chatFile_row, column=chatFile_col).value is not None:

        firstValue = chatFile_sheet.cell(row=chatFile_row, column=chatFile_col).value
        secondValue = chatFile_sheet.cell(row=chatFile_row, column=chatFile_col+1).value
        return firstValue, secondValue

    else:
        return None
